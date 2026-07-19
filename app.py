import streamlit as st
import cv2
import numpy as np
import pytesseract
import pandas as pd
import tempfile
import os
import sys
import shutil
import io
import time
from PIL import Image, ImageEnhance, ImageFilter
from streamlit_drawable_canvas import st_canvas
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from fpdf import FPDF

# ─────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────
st.set_page_config(
    page_title="Video OCR Extractor",
    page_icon="🎬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────
# TESSERACT AUTO-DETECTION (local / packaged .exe)
# ─────────────────────────────────────────
def _app_base_dir() -> str:
    """Directory of the running app, whether as a script or a frozen .exe."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def configure_tesseract() -> bool:
    """Locate the Tesseract binary for local/packaged runs.

    Order: TESSERACT_CMD env var -> a "tesseract" folder shipped next to the
    exe -> the default Windows install path -> whatever is already on PATH.
    Returns True if a usable tesseract was found.
    """
    custom = os.environ.get("TESSERACT_CMD")
    if custom and os.path.isfile(custom):
        pytesseract.pytesseract.tesseract_cmd = custom
        return True

    base = _app_base_dir()
    bundled = os.path.join(base, "tesseract", "tesseract.exe")
    if os.path.isfile(bundled):
        pytesseract.pytesseract.tesseract_cmd = bundled
        return True

    default_win = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if os.name == "nt" and os.path.isfile(default_win):
        pytesseract.pytesseract.tesseract_cmd = default_win
        return True

    return shutil.which(pytesseract.pytesseract.tesseract_cmd or "tesseract") is not None


TESSERACT_OK = configure_tesseract()

# ─────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&display=swap');

:root {
    --acc: #00d4ff;
    --acc2: #ff6b35;
    --bg: #0f1117;
    --surf: #161920;
    --grn: #00e5a0;
}

html, body, [class*="css"] {
    font-family: 'Noto Sans KR', sans-serif;
}

/* Header */
.app-header {
    background: linear-gradient(135deg, #0f1117 0%, #161920 100%);
    border: 1px solid #252830;
    border-radius: 12px;
    padding: 20px 28px;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    gap: 16px;
}
.app-title {
    font-family: 'JetBrains Mono', monospace;
    font-size: 22px;
    font-weight: 700;
    letter-spacing: 0.08em;
    background: linear-gradient(135deg, #00d4ff, #ff6b35);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    margin: 0;
}
.app-sub {
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px;
    color: #5a6072;
    margin: 0;
}

/* Metric cards */
.metric-card {
    background: #161920;
    border: 1px solid #252830;
    border-radius: 8px;
    padding: 12px 16px;
    text-align: center;
}
.metric-val {
    font-family: 'JetBrains Mono', monospace;
    font-size: 28px;
    font-weight: 700;
    color: #00d4ff;
}
.metric-lbl {
    font-size: 11px;
    color: #5a6072;
    margin-top: 2px;
}

/* Step badge */
.step-badge {
    background: linear-gradient(135deg, rgba(0,212,255,0.15), rgba(255,107,53,0.1));
    border: 1px solid rgba(0,212,255,0.3);
    border-radius: 6px;
    padding: 6px 12px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px;
    color: #00d4ff;
    letter-spacing: 0.1em;
    display: inline-block;
    margin-bottom: 8px;
}

/* Result table styling */
.stDataFrame { border: 1px solid #252830 !important; border-radius: 8px !important; }

/* Download buttons */
.dl-btn-row { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 10px; }

/* Confidence pill */
.conf-hi { color: #00e5a0; font-weight: 600; }
.conf-md { color: #ffd166; font-weight: 600; }
.conf-lo { color: #ff4757; font-weight: 600; }

/* Info box */
.info-box {
    background: rgba(0,212,255,0.06);
    border: 1px solid rgba(0,212,255,0.2);
    border-radius: 8px;
    padding: 12px 16px;
    font-size: 13px;
    color: #adb5c8;
    margin: 8px 0;
}

/* Placeholder box (no external image needed -> works fully offline) */
.video-placeholder {
    display: flex;
    align-items: center;
    justify-content: center;
    flex-direction: column;
    gap: 10px;
    height: 340px;
    background: #0f1117;
    border: 1px dashed #252830;
    border-radius: 10px;
    color: #5a6072;
    font-family: 'JetBrains Mono', monospace;
    font-size: 13px;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <div style="font-size:36px">🎬</div>
  <div>
    <p class="app-title">VIDEO OCR EXTRACTOR</p>
    <p class="app-sub">영역 선택 → 프레임 캡처 → 숫자/텍스트 추출 → 데이터 저장</p>
  </div>
  <div style="margin-left:auto;font-family:'JetBrains Mono',monospace;font-size:11px;
              color:#00d4ff;border:1px solid #00d4ff;border-radius:4px;padding:3px 10px;">v2.1</div>
</div>
""", unsafe_allow_html=True)

if not TESSERACT_OK:
    st.error(
        "⚠️ Tesseract OCR 엔진을 찾을 수 없습니다. "
        "[Tesseract-OCR for Windows](https://github.com/UB-Mannheim/tesseract/wiki)를 설치하거나, "
        "`tesseract` 폴더를 실행 파일 옆에 두거나, `TESSERACT_CMD` 환경변수로 경로를 지정하세요."
    )

# ─────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────
for key, val in {
    'results': [],
    'video_path': None,
    'total_frames': 0,
    'fps': 0,
    'duration': 0,
    'frame_preview': None,
    'region': None,
    'running': False,
}.items():
    if key not in st.session_state:
        st.session_state[key] = val


# ─────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────
def fmt_time(seconds: float) -> str:
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    ms = int((seconds % 1) * 1000)
    if h > 0:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}.{ms:03d}"


def extract_frame(cap, time_sec: float) -> np.ndarray | None:
    cap.set(cv2.CAP_PROP_POS_MSEC, time_sec * 1000)
    ret, frame = cap.read()
    return frame if ret else None


def preprocess(img: np.ndarray, mode: str) -> np.ndarray:
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    if mode == "inv":
        gray = cv2.bitwise_not(gray)
    if mode in ("auto", "inv"):
        gray = cv2.convertScaleAbs(gray, alpha=1.8, beta=-30)
    # CLAHE for local contrast
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(4, 4))
    gray = clahe.apply(gray)
    return gray


def run_ocr(img_gray: np.ndarray, lang: str, mode: str, scale: int) -> tuple[str, int]:
    h, w = img_gray.shape
    big = cv2.resize(img_gray, (w * scale, h * scale), interpolation=cv2.INTER_CUBIC)

    psm = "6"
    if mode in ("num", "dig"):
        psm = "7" if h * scale < 60 else "6"
    cfg = f"--psm {psm} --oem 3"
    if mode == "dig":
        cfg += " -c tessedit_char_whitelist=0123456789."
    elif mode == "num":
        cfg += " -c tessedit_char_whitelist=0123456789.-+eE/ "

    pil = Image.fromarray(big)
    data = pytesseract.image_to_data(pil, lang=lang, config=cfg, output_type=pytesseract.Output.DICT)

    words, confs = [], []
    for i, conf in enumerate(data['conf']):
        try:
            c = int(conf)
        except Exception:
            continue
        if c > 0:
            words.append(data['text'][i])
            confs.append(c)

    raw = ' '.join(words).strip()
    avg_conf = int(sum(confs) / len(confs)) if confs else 0

    # Post-process filter
    if mode == "dig":
        lines = [l.replace(' ', '') for l in raw.split('\n')]
        raw = '\n'.join(l for l in lines if any(c.isdigit() for c in l))
    elif mode == "num":
        import re
        parts = re.findall(r'[\d.+\-eE/]+', raw)
        raw = '  '.join(p for p in parts if any(c.isdigit() for c in p))

    return raw.strip(), avg_conf


def conf_label(c: int) -> str:
    if c >= 70:
        return "🟢 높음"
    elif c >= 40:
        return "🟡 보통"
    return "🔴 낮음"


# ─────────────────────────────────────────
# SIDEBAR — SETTINGS
# ─────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="step-badge">⚙️ SETTINGS</div>', unsafe_allow_html=True)

    # ── Video upload
    st.markdown("**📁 비디오 파일**")
    uploaded = st.file_uploader(
        "MP4 · MOV · AVI · WebM",
        type=["mp4", "mov", "avi", "webm", "mkv"],
        label_visibility="collapsed",
    )

    if uploaded:
        # Save to temp file
        suffix = os.path.splitext(uploaded.name)[1]
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        tmp.write(uploaded.read())
        tmp.flush()
        st.session_state.video_path = tmp.name

        cap = cv2.VideoCapture(st.session_state.video_path)
        st.session_state.fps      = cap.get(cv2.CAP_PROP_FPS) or 30
        st.session_state.total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        st.session_state.duration = st.session_state.total_frames / st.session_state.fps
        cap.release()

        st.success(f"✓ {uploaded.name}")
        st.caption(f"길이: {fmt_time(st.session_state.duration)}  |  FPS: {st.session_state.fps:.1f}")

    st.divider()

    # ── Extraction settings
    st.markdown("**⚙️ 추출 설정**")
    col1, col2 = st.columns(2)
    with col1:
        interval = st.number_input("간격 (초)", min_value=0.1, max_value=60.0, value=1.0, step=0.1, format="%.1f")
    with col2:
        lang_map = {"영어 (숫자 권장)": "eng", "한국어+영어": "kor+eng", "한국어": "kor"}
        lang_label = st.selectbox("언어", list(lang_map.keys()))
        lang = lang_map[lang_label]

    col3, col4 = st.columns(2)
    with col3:
        t_start = st.number_input("시작 (초)", min_value=0.0, value=0.0, step=1.0)
    with col4:
        dur = st.session_state.duration
        t_end = st.number_input("종료 (초)", min_value=0.0,
                                value=float(int(dur)) if dur > 0 else 0.0,
                                step=1.0,
                                help="0이면 전체")

    col5, col6 = st.columns(2)
    with col5:
        mode_map = {"숫자 위주": "num", "숫자만": "dig", "일반 텍스트": "all"}
        mode_label = st.selectbox("OCR 모드", list(mode_map.keys()))
        mode = mode_map[mode_label]
    with col6:
        scale = st.selectbox("이미지 확대", [3, 2, 4, 1], format_func=lambda x: f"{x}× {'(권장)' if x==3 else ''}")

    prep_map = {"자동 (그레이+대비)": "auto", "반전+대비 (밝은 LCD)": "inv", "원본 유지": "none"}
    prep_label = st.selectbox("이미지 전처리", list(prep_map.keys()))
    prep = prep_map[prep_label]

    st.divider()

    # ── Preview frame slider
    st.markdown("**🎞️ 미리보기 프레임**")
    if st.session_state.video_path and st.session_state.duration > 0:
        preview_sec = st.slider(
            "시간 (초)",
            min_value=0.0,
            max_value=float(int(st.session_state.duration)),
            value=0.0,
            step=1.0,
            label_visibility="collapsed",
        )
        cap = cv2.VideoCapture(st.session_state.video_path)
        f = extract_frame(cap, preview_sec)
        cap.release()
        if f is not None:
            st.session_state.frame_preview = cv2.cvtColor(f, cv2.COLOR_BGR2RGB)
    else:
        preview_sec = 0.0

    st.divider()
    st.markdown("**ℹ️ 사용법**")
    st.markdown("""
<div class="info-box">
1. 비디오 업로드<br>
2. 오른쪽 캔버스에서 OCR 영역 드래그<br>
3. 설정 후 <b>추출 시작</b> 클릭<br>
4. 결과를 CSV / Excel / PDF로 저장
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────
# MAIN AREA
# ─────────────────────────────────────────
left_col, right_col = st.columns([1.1, 1], gap="large")

with left_col:
    # ── STEP 1: Region selection
    st.markdown('<div class="step-badge">STEP 1 — OCR 영역 선택</div>', unsafe_allow_html=True)

    if st.session_state.frame_preview is not None:
        frame_img = st.session_state.frame_preview
        fh, fw = frame_img.shape[:2]
        # Scale for display (max width 600)
        disp_w = 600
        disp_h = int(fh * disp_w / fw)

        canvas_result = st_canvas(
            fill_color="rgba(0, 212, 255, 0.08)",
            stroke_width=2,
            stroke_color="#00d4ff",
            background_image=Image.fromarray(frame_img),
            update_streamlit=True,
            width=disp_w,
            height=disp_h,
            drawing_mode="rect",
            key="canvas",
        )

        # Parse region
        if (canvas_result.json_data is not None and
                canvas_result.json_data.get("objects")):
            objs = canvas_result.json_data["objects"]
            if objs:
                last = objs[-1]
                scale_x = fw / disp_w
                scale_y = fh / disp_h
                rx = int(last.get("left", 0) * scale_x)
                ry = int(last.get("top", 0) * scale_y)
                rw = int(last.get("width", 0) * scale_x)
                rh = int(last.get("height", 0) * scale_y)
                if rw > 10 and rh > 10:
                    st.session_state.region = (rx, ry, rw, rh)

        # Region info
        r = st.session_state.region
        if r:
            c1, c2, c3, c4 = st.columns(4)
            for col, lbl, val in zip([c1,c2,c3,c4], ["X","Y","W","H"], r):
                with col:
                    st.markdown(f"""<div class="metric-card">
                        <div class="metric-val">{val}</div>
                        <div class="metric-lbl">{lbl}</div></div>""",
                        unsafe_allow_html=True)

            # Preview crop
            st.markdown("**🔍 영역 미리보기 (전처리 적용)**")
            frame_bgr = cv2.cvtColor(frame_img, cv2.COLOR_RGB2BGR)
            crop = frame_bgr[r[1]:r[1]+r[3], r[0]:r[0]+r[2]]
            if crop.size > 0:
                gray = preprocess(crop, prep)
                bh, bw = gray.shape
                big = cv2.resize(gray, (bw * int(scale), bh * int(scale)), interpolation=cv2.INTER_CUBIC)
                st.image(big, caption=f"전처리 후 ({bw*scale}×{bh*scale}px)", use_column_width=True)
        else:
            st.markdown('<div class="info-box">↑ 캔버스에서 OCR 영역을 드래그하여 선택하세요</div>',
                        unsafe_allow_html=True)
    else:
        st.markdown('<div class="info-box">← 비디오를 먼저 업로드하세요</div>',
                    unsafe_allow_html=True)
        st.markdown(
            '<div class="video-placeholder">🎬<br>Video Preview</div>',
            unsafe_allow_html=True,
        )


with right_col:
    # ── STEP 2: Run OCR
    st.markdown('<div class="step-badge">STEP 2 — OCR 추출 실행</div>', unsafe_allow_html=True)

    can_run = (
        st.session_state.video_path is not None and
        st.session_state.duration > 0 and
        TESSERACT_OK
    )

    col_run, col_clear = st.columns([2, 1])
    with col_run:
        run_btn = st.button(
            "▶ OCR 추출 시작",
            disabled=not can_run,
            use_container_width=True,
            type="primary",
        )
    with col_clear:
        if st.button("🗑 결과 초기화", use_container_width=True):
            st.session_state.results = []
            st.rerun()

    # ── OCR LOOP ──────────────────────────────
    if run_btn and can_run:
        end_sec = t_end if t_end > 0 else st.session_state.duration
        if t_start >= end_sec:
            st.error("시작 시간이 종료 시간보다 크거나 같습니다.")
        else:
            st.session_state.results = []
            timestamps = []
            t = t_start
            while t <= end_sec + 1e-6:
                timestamps.append(min(t, end_sec))
                t = round(t + interval, 6)

            total = len(timestamps)
            prog_bar  = st.progress(0, text="초기화 중...")
            stat_text = st.empty()
            result_placeholder = st.empty()

            cap = cv2.VideoCapture(st.session_state.video_path)
            r = st.session_state.region

            for i, ts in enumerate(timestamps):
                pct = int(i / total * 100)
                prog_bar.progress(pct / 100, text=f"{i+1}/{total} 프레임  |  {fmt_time(ts)}")
                stat_text.caption(f"처리 중: {fmt_time(ts)}")

                frame = extract_frame(cap, ts)
                if frame is None:
                    st.session_state.results.append({
                        "타임코드": fmt_time(ts),
                        "시간(초)": round(ts, 3),
                        "신뢰도(%)": 0,
                        "신뢰도 등급": "오류",
                        "추출값": "(프레임 추출 실패)",
                    })
                    continue

                # Crop region
                if r and r[2] > 0 and r[3] > 0:
                    fh, fw = frame.shape[:2]
                    x1 = max(0, r[0]); y1 = max(0, r[1])
                    x2 = min(fw, r[0]+r[2]); y2 = min(fh, r[1]+r[3])
                    crop = frame[y1:y2, x1:x2]
                else:
                    crop = frame

                if crop.size == 0:
                    st.session_state.results.append({
                        "타임코드": fmt_time(ts),
                        "시간(초)": round(ts, 3),
                        "신뢰도(%)": 0,
                        "신뢰도 등급": "오류",
                        "추출값": "(영역 오류)",
                    })
                    continue

                # Preprocess + OCR
                gray = preprocess(crop, prep)
                ocr_lang = "eng" if mode in ("num", "dig") else lang
                text, conf = run_ocr(gray, ocr_lang, mode, int(scale))

                st.session_state.results.append({
                    "타임코드": fmt_time(ts),
                    "시간(초)": round(ts, 3),
                    "신뢰도(%)": conf,
                    "신뢰도 등급": conf_label(conf),
                    "추출값": text if text else "(없음)",
                })

                # Live table update every 5 rows
                if i % 5 == 0 and st.session_state.results:
                    df_live = pd.DataFrame(st.session_state.results[-20:])
                    result_placeholder.dataframe(df_live[["타임코드","신뢰도(%)","추출값"]], use_container_width=True)

            cap.release()
            prog_bar.progress(1.0, text=f"완료 — {len(st.session_state.results)}건 추출 ✓")
            stat_text.empty()
            st.success(f"✅ OCR 추출 완료: **{len(st.session_state.results)}건**")

    # ── RESULTS ──────────────────────────────
    st.markdown("---")
    st.markdown('<div class="step-badge">STEP 3 — 결과 확인 및 저장</div>', unsafe_allow_html=True)

    results = st.session_state.results
    if results:
        df = pd.DataFrame(results)

        # Summary metrics
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f"""<div class="metric-card">
                <div class="metric-val">{len(df)}</div>
                <div class="metric-lbl">총 추출 건수</div></div>""", unsafe_allow_html=True)
        with c2:
            avg = int(df["신뢰도(%)"].mean())
            st.markdown(f"""<div class="metric-card">
                <div class="metric-val">{avg}%</div>
                <div class="metric-lbl">평균 신뢰도</div></div>""", unsafe_allow_html=True)
        with c3:
            hi = int((df["신뢰도(%)"] >= 70).sum())
            st.markdown(f"""<div class="metric-card">
                <div class="metric-val">{hi}</div>
                <div class="metric-lbl">높음 (≥70%)</div></div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.dataframe(df, use_container_width=True, height=340)

        # ── EXPORT BUTTONS ────────────────────
        st.markdown("**📦 내보내기**")
        dl1, dl2, dl3, dl4 = st.columns(4)

        # CSV
        with dl1:
            csv_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button(
                "📄 CSV 저장",
                data=csv_bytes,
                file_name=f"ocr_{int(time.time())}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # Excel
        with dl2:
            excel_buf = io.BytesIO()
            with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
                # Sheet 1: Raw data
                df.to_excel(writer, index=False, sheet_name="OCR 데이터")
                ws = writer.sheets["OCR 데이터"]

                # Style header
                hdr_fill = PatternFill("solid", fgColor="0F1117")
                acc_fill = PatternFill("solid", fgColor="00D4FF")
                hdr_font = Font(bold=True, color="000000", name="Arial", size=10)
                for cell in ws[1]:
                    cell.fill = acc_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Col widths
                for col, w in zip(ws.columns, [16, 12, 12, 14, 40]):
                    ws.column_dimensions[col[0].column_letter].width = w

                # Zebra rows
                alt_fill = PatternFill("solid", fgColor="161920")
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if i % 2 == 0:
                        for cell in row:
                            cell.fill = alt_fill

                # Sheet 2: Pivot (numbers split)
                pivot = []
                for row in results:
                    vals = [v.strip() for v in row["추출값"].replace('/', '\n').split('\n') if v.strip()]
                    pivot.append({
                        "타임코드": row["타임코드"],
                        "시간(초)": row["시간(초)"],
                        "값1": vals[0] if len(vals) > 0 else "",
                        "값2": vals[1] if len(vals) > 1 else "",
                        "값3": vals[2] if len(vals) > 2 else "",
                    })
                pd.DataFrame(pivot).to_excel(writer, index=False, sheet_name="숫자 피벗")
                ws2 = writer.sheets["숫자 피벗"]
                for cell in ws2[1]:
                    cell.fill = acc_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center")
                for col in ws2.columns:
                    ws2.column_dimensions[col[0].column_letter].width = 16

            st.download_button(
                "📊 Excel 저장",
                data=excel_buf.getvalue(),
                file_name=f"ocr_{int(time.time())}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # PDF
        with dl3:
            class PDF(FPDF):
                def header(self):
                    self.set_fill_color(15, 17, 23)
                    self.rect(0, 0, 297, 210, 'F')
                    self.set_font("Helvetica", "B", 14)
                    self.set_text_color(0, 212, 255)
                    self.cell(0, 10, "VIDEO OCR EXTRACTOR - Results", ln=True)
                    self.set_font("Helvetica", "", 8)
                    self.set_text_color(90, 96, 114)
                    self.cell(0, 6, f"Total: {len(results)} rows  |  Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
                    self.ln(2)

                def footer(self):
                    self.set_y(-12)
                    self.set_font("Helvetica", "", 7)
                    self.set_text_color(90, 96, 114)
                    self.cell(0, 10, f"Page {self.page_no()}", align="R")

            pdf = PDF(orientation="L", unit="mm", format="A4")
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=12)

            # Table header
            cols = ["타임코드", "시간(초)", "신뢰도(%)", "추출값"]
            widths = [35, 28, 25, 190]
            pdf.set_fill_color(0, 212, 255)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "B", 8)
            for col, w in zip(cols, widths):
                pdf.cell(w, 8, col, border=1, align="C", fill=True)
            pdf.ln()

            # Rows
            for i, row in enumerate(results):
                fill = i % 2 == 0
                pdf.set_fill_color(22, 25, 32) if fill else pdf.set_fill_color(15, 17, 23)
                pdf.set_text_color(221, 225, 236)
                pdf.set_font("Helvetica", "", 7)
                vals = [
                    row["타임코드"],
                    str(row["시간(초)"]),
                    f"{row['신뢰도(%)']}%",
                    row["추출값"].replace('\n', ' / ')[:80],
                ]
                for val, w in zip(vals, widths):
                    pdf.cell(w, 6, val, border=1, fill=fill)
                pdf.ln()

            pdf_bytes = pdf.output(dest="S").encode("latin-1")
            st.download_button(
                "📋 PDF 저장",
                data=pdf_bytes,
                file_name=f"ocr_{int(time.time())}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        # Clipboard / TSV
        with dl4:
            tsv = "\n".join(
                f"[{r['타임코드']}]\t{r['추출값'].replace(chr(10),' ')}"
                for r in results
            )
            st.download_button(
                "📎 TSV 저장",
                data=tsv.encode("utf-8-sig"),
                file_name=f"ocr_{int(time.time())}.tsv",
                mime="text/tab-separated-values",
                use_container_width=True,
            )

    else:
        st.markdown("""
<div class="info-box" style="text-align:center;padding:40px">
  <div style="font-size:48px;opacity:.15;margin-bottom:12px">📋</div>
  OCR 추출 결과가 여기에 표시됩니다
</div>
""", unsafe_allow_html=True)
